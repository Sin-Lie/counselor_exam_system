# -*- coding: utf-8 -*-
# Excel导入导出工具类
# 封装 openpyxl 操作，供 system/services.py 和 score/services.py 复用

import io
import openpyxl
from urllib.parse import quote
from django.http import HttpResponse


def read_excel(file_obj):
    """
    读取 Excel 文件，返回工作表对象和所有数据行
    返回: (worksheet, rows) — rows 为 list[tuple]，每行为单元格值的元组
    """
    wb = openpyxl.load_workbook(file_obj)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    return ws, rows


def read_excel_with_headers(file_obj):
    """
    读取 Excel 文件，返回表头列表和数据行
    返回: (headers, rows)
        headers: list[str] — 第一行表头文本列表
        rows: list[tuple] — 从第二行开始的数据行
    """
    wb = openpyxl.load_workbook(file_obj)
    ws = wb.active
    headers = [str(cell.value).strip() if cell.value is not None else '' for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    return headers, rows


def build_column_map(headers, field_mapping):
    """
    根据表头行构建 字段名 → 列索引 的映射（不区分大小写）
    参数:
        headers: list[str] — Excel 表头
        field_mapping: dict[str, list[str]] — 字段名 → 可能的表头名称列表
    返回: dict[str, int] — 字段名 → 列索引，未匹配到的字段不存在于返回中
    """
    # 对表头做标准化：去空格、转小写，建立 标准化名称 → 索引 的映射
    normalized = {}
    for idx, h in enumerate(headers):
        key = h.strip().lower().replace('（', '(').replace('）', ')')
        # 如果标准化后有重复，保留第一个出现的索引
        if key not in normalized:
            normalized[key] = idx

    result = {}
    for field, aliases in field_mapping.items():
        for alias in aliases:
            alias_key = alias.strip().lower().replace('（', '(').replace('）', ')')
            if alias_key in normalized:
                result[field] = normalized[alias_key]
                break

    return result


def safe_cell_str(row, idx):
    """安全获取单元格字符串值，若 idx 为负数（列不存在）则返回空字符串"""
    if idx < 0:
        return ''
    if len(row) > idx and row[idx] is not None:
        return str(row[idx]).strip()
    return ''


def create_excel_response(filename, sheet_title='Sheet1', headers=None, rows=None):
    """
    创建 Excel 文件并返回 HttpResponse 供下载
    参数:
        filename: 下载文件名（不含 .xlsx 后缀）
        sheet_title: 工作表名称
        headers: 表头列表
        rows: 数据行列表（每行为 list/tuple）
    返回: HttpResponse
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title

    if headers:
        ws.append(headers)

    if rows:
        for row in rows:
            ws.append(row)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    safe_name = quote(f"{filename}.xlsx")
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f"attachment; filename*=UTF-8''{safe_name}"
    return response
